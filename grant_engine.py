"""
grant_engine.py — All grant scanning, DB, keyword, and export logic.
Consolidated from the Jupyter notebook into a single importable module.
Includes full two-sheet Excel export with Calendar View.
"""

import requests
import feedparser
import sqlite3
import hashlib
import json
import re
import time
import logging
from bs4 import BeautifulSoup
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
from contextlib import contextmanager
import pandas as pd
from datetime import datetime, timedelta, timezone

try:
    from dateutil.relativedelta import relativedelta
    HAS_DATEUTIL = True
except ImportError:
    HAS_DATEUTIL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ==============================================================
# UTILITIES
# ==============================================================

def utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("grant_monitor.log"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("grant_monitor")

CONFIG = {
    "db_path": "grants.db",
    "request_delay_seconds": 2,
    "request_timeout_seconds": 30,
    "user_agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
}

DEFAULT_KEYWORDS = [
    "cybersecurity", "AI", "experiential learning", "medical technology",
    "data science", "machine learning", "artificial intelligence",
    "STEM education", "novel education", "workforce",
    "workforce development", "quantum",
]


# ==============================================================
# DATABASE
# ==============================================================

@contextmanager
def get_db(db_path: str = None):
    if db_path is None:
        db_path = CONFIG["db_path"]
    conn = sqlite3.connect(db_path, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


@dataclass
class Grant:
    title: str
    url: str
    source: str
    description: str = ""
    deadline: Optional[str] = None
    posted_date: Optional[str] = None
    funding_amount: Optional[str] = None
    matched_keywords: list = field(default_factory=list)
    content_hash: str = ""
    first_seen: str = ""
    last_seen: str = ""
    is_new: bool = True

    def compute_hash(self):
        raw = f"{self.title}|{self.description}|{self.deadline}".lower().strip()
        self.content_hash = hashlib.sha256(raw.encode()).hexdigest()
        return self.content_hash


def init_db(db_path: str = None):
    if db_path is None:
        db_path = CONFIG["db_path"]
    with get_db(db_path) as conn:
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS grants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL, url TEXT NOT NULL, source TEXT NOT NULL,
                description TEXT, deadline TEXT, posted_date TEXT,
                funding_amount TEXT, matched_keywords TEXT,
                content_hash TEXT UNIQUE, first_seen TEXT NOT NULL,
                last_seen TEXT NOT NULL, notified INTEGER DEFAULT 0
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS scan_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source TEXT NOT NULL, scan_time TEXT NOT NULL,
                grants_found INTEGER DEFAULT 0, new_grants INTEGER DEFAULT 0,
                updated_grants INTEGER DEFAULT 0, status TEXT DEFAULT 'success',
                error_message TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT UNIQUE NOT NULL,
                added_date TEXT NOT NULL, active INTEGER DEFAULT 1
            )
        """)
        now = utc_now().strftime("%Y-%m-%dT%H:%M:%S")
        for kw in DEFAULT_KEYWORDS:
            c.execute(
                "INSERT OR IGNORE INTO keywords (keyword, added_date) VALUES (?, ?)",
                (kw, now),
            )
    logger.info("Database initialized.")


def upsert_grant(grant: Grant, db_path: str = None) -> str:
    if db_path is None:
        db_path = CONFIG["db_path"]
    now = utc_now().strftime("%Y-%m-%dT%H:%M:%S")
    grant.compute_hash()
    grant.last_seen = now

    with get_db(db_path) as conn:
        c = conn.cursor()
        c.execute("SELECT id, content_hash FROM grants WHERE url = ?", (grant.url,))
        row = c.fetchone()

        if row is None:
            grant.first_seen = now
            c.execute("""
                INSERT INTO grants
                (title, url, source, description, deadline, posted_date,
                 funding_amount, matched_keywords, content_hash, first_seen, last_seen)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                grant.title, grant.url, grant.source, grant.description,
                grant.deadline, grant.posted_date, grant.funding_amount,
                json.dumps(grant.matched_keywords), grant.content_hash,
                grant.first_seen, grant.last_seen,
            ))
            return "new"
        elif row[1] != grant.content_hash:
            c.execute("""
                UPDATE grants SET title=?, description=?, deadline=?, posted_date=?,
                    funding_amount=?, matched_keywords=?, content_hash=?,
                    last_seen=?, notified=0
                WHERE url=?
            """, (
                grant.title, grant.description, grant.deadline,
                grant.posted_date, grant.funding_amount,
                json.dumps(grant.matched_keywords), grant.content_hash,
                grant.last_seen, grant.url,
            ))
            return "updated"
        else:
            c.execute("UPDATE grants SET last_seen=? WHERE url=?", (now, grant.url))
            return "unchanged"


def log_scan(source, grants_found, new_grants, updated_grants,
             status="success", error_message=None):
    with get_db() as conn:
        conn.execute("""
            INSERT INTO scan_log
            (source, scan_time, grants_found, new_grants, updated_grants, status, error_message)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (source, utc_now().strftime("%Y-%m-%dT%H:%M:%S"), grants_found,
              new_grants, updated_grants, status, error_message))


# ==============================================================
# KEYWORDS
# ==============================================================

def get_active_keywords(db_path: str = None) -> list:
    if db_path is None:
        db_path = CONFIG["db_path"]
    with get_db(db_path) as conn:
        c = conn.cursor()
        c.execute("SELECT keyword FROM keywords WHERE active = 1")
        return [row[0] for row in c.fetchall()]


def add_keyword(kw: str):
    now = utc_now().strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO keywords (keyword, added_date, active) VALUES (?, ?, 1)",
            (kw, now),
        )
        conn.execute(
            "UPDATE keywords SET active = 1 WHERE keyword = ? AND active = 0", (kw,)
        )


def remove_keyword(kw: str):
    with get_db() as conn:
        conn.execute("UPDATE keywords SET active = 0 WHERE keyword = ?", (kw,))


def match_keywords(text: str, keywords: list = None) -> list:
    if keywords is None:
        keywords = get_active_keywords()
    text_lower = text.lower()
    matched = []
    for kw in keywords:
        if kw == "AI":
            if re.search(r'\bAI\b', text):
                matched.append(kw)
        else:
            if re.search(r'\b' + re.escape(kw.lower()) + r'\b', text_lower):
                matched.append(kw)
    return matched


def passes_filter(title: str, description: str) -> list:
    return match_keywords(f"{title} {description}")


# ==============================================================
# HTTP SESSION
# ==============================================================

def get_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": CONFIG["user_agent"],
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


SESSION = get_session()


def safe_request(url: str, **kwargs):
    try:
        time.sleep(CONFIG["request_delay_seconds"])
        resp = SESSION.get(url, timeout=CONFIG["request_timeout_seconds"], **kwargs)
        resp.raise_for_status()
        return resp
    except requests.RequestException as e:
        logger.error(f"Request failed for {url}: {e}")
        return None


# ==============================================================
# SCRAPERS
# ==============================================================

def scrape_grants_gov(keywords=None):
    if keywords is None:
        keywords = get_active_keywords()
    grants = []
    base_url = "https://apply07.grants.gov/grantsws/rest/opportunities/search/"
    for kw in keywords:
        payload = {
            "keyword": kw, "oppStatuses": "forecasted|posted",
            "sortBy": "openDate|desc", "rows": 50, "startRecordNum": 0,
        }
        try:
            time.sleep(CONFIG["request_delay_seconds"])
            resp = requests.post(base_url, json=payload,
                                 headers={"Content-Type": "application/json"},
                                 timeout=CONFIG["request_timeout_seconds"])
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            logger.error(f"Grants.gov API error for '{kw}': {e}")
            continue
        opportunities = data.get("oppHits", [])
        logger.info(f"Grants.gov: '{kw}' returned {len(opportunities)} results")
        for opp in opportunities:
            title = opp.get("title", "")
            description = opp.get("description", title)
            url = f"https://www.grants.gov/search-results-detail/{opp.get('id', '')}"
            matched = passes_filter(title, description)
            if matched:
                g = Grant(
                    title=title, url=url, source="grants.gov",
                    description=description[:2000],
                    deadline=opp.get("closeDate", ""),
                    posted_date=opp.get("openDate", ""),
                    funding_amount=opp.get("awardCeiling", ""),
                    matched_keywords=matched,
                )
                grants.append(g)
    seen = set()
    unique = []
    for g in grants:
        if g.url not in seen:
            seen.add(g.url)
            unique.append(g)
    logger.info(f"Grants.gov total unique matches: {len(unique)}")
    return unique


def scrape_nsf():
    feeds = [
        "https://www.nsf.gov/rss/rss_www_funding.xml",
        "https://www.nsf.gov/rss/rss_www_funding_pgm_annc_cise.xml",
        "https://www.nsf.gov/rss/rss_www_funding_pgm_annc_ehr.xml",
    ]
    grants = []
    for feed_url in feeds:
        try:
            time.sleep(CONFIG["request_delay_seconds"])
            feed = feedparser.parse(feed_url)
            logger.info(f"NSF feed: {len(feed.entries)} entries from {feed_url}")
        except Exception as e:
            logger.error(f"NSF feed error ({feed_url}): {e}")
            continue
        for entry in feed.entries:
            title = entry.get("title", "")
            summary = entry.get("summary", "")
            link = entry.get("link", "")
            published = entry.get("published", "")
            matched = passes_filter(title, summary)
            if matched:
                g = Grant(title=title, url=link, source="nsf.gov",
                          description=summary[:2000], posted_date=published,
                          matched_keywords=matched)
                grants.append(g)
    seen = set()
    unique = []
    for g in grants:
        if g.url not in seen:
            seen.add(g.url)
            unique.append(g)
    logger.info(f"NSF total unique matches: {len(unique)}")
    return unique


def scrape_nih_reporter():
    grants = []
    keywords = get_active_keywords()
    url = "https://api.reporter.nih.gov/v2/projects/search"
    search_text = " OR ".join(
        [kw for kw in keywords if kw.lower() in
         ("medical technology", "ai", "artificial intelligence",
          "data science", "machine learning", "cybersecurity")]
    )
    payload = {
        "criteria": {
            "advanced_text_search": {
                "operator": "or", "search_field": "projecttitle,terms",
                "search_text": search_text,
            },
            "fiscal_years": [2025, 2026],
            "newly_added_projects_only": True,
        },
        "offset": 0, "limit": 50,
        "sort_field": "project_start_date", "sort_order": "desc",
    }
    try:
        time.sleep(CONFIG["request_delay_seconds"])
        resp = requests.post(url, json=payload, timeout=CONFIG["request_timeout_seconds"])
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        logger.error(f"NIH Reporter API error: {e}")
        return grants
    results = data.get("results", [])
    logger.info(f"NIH Reporter: {len(results)} results")
    for proj in results:
        title = proj.get("project_title", "")
        abstract = proj.get("abstract_text", "") or ""
        link = f"https://reporter.nih.gov/project-details/{proj.get('application_id', '')}"
        matched = passes_filter(title, abstract)
        if matched:
            award = proj.get("award_amount")
            g = Grant(
                title=title.strip(), url=link, source="nih_reporter",
                description=abstract[:2000],
                posted_date=proj.get("project_start_date", ""),
                funding_amount=str(award) if award else "",
                matched_keywords=matched,
            )
            grants.append(g)
    logger.info(f"NIH Reporter matches: {len(grants)}")
    return grants


def scrape_rss_source(feed_url, source_name):
    grants = []
    try:
        time.sleep(CONFIG["request_delay_seconds"])
        feed = feedparser.parse(feed_url)
        logger.info(f"{source_name}: {len(feed.entries)} entries")
    except Exception as e:
        logger.error(f"{source_name} feed error: {e}")
        return grants
    for entry in feed.entries:
        title = entry.get("title", "")
        summary = entry.get("summary", entry.get("description", ""))
        link = entry.get("link", "")
        published = entry.get("published", entry.get("updated", ""))
        matched = passes_filter(title, summary)
        if matched:
            g = Grant(title=title, url=link, source=source_name,
                      description=summary[:2000], posted_date=published,
                      matched_keywords=matched)
            grants.append(g)
    logger.info(f"{source_name} matches: {len(grants)}")
    return grants


ADDITIONAL_RSS_SOURCES = [
    {"url": "https://www.ed.gov/feed", "name": "ed.gov"},
    {"url": "https://www.sbir.gov/rss-feed", "name": "sbir.gov"},
    {"url": "https://www.darpa.mil/rss", "name": "darpa.mil"},
]


# ==============================================================
# FULL SCAN
# ==============================================================

def run_full_scan() -> dict:
    logger.info("=" * 60)
    logger.info("STARTING FULL GRANT SCAN")
    logger.info(f"Active keywords: {get_active_keywords()}")
    logger.info("=" * 60)

    summary = {
        "scan_time": utc_now().isoformat(),
        "sources_scanned": 0, "total_found": 0,
        "new": 0, "updated": 0, "unchanged": 0, "errors": [],
    }
    all_grants = []

    try:
        results = scrape_grants_gov()
        all_grants.extend(results)
        log_scan("grants.gov", len(results), 0, 0)
        summary["sources_scanned"] += 1
    except Exception as e:
        logger.error(f"Grants.gov scraper failed: {e}")
        log_scan("grants.gov", 0, 0, 0, "error", str(e))
        summary["errors"].append(f"grants.gov: {e}")

    try:
        results = scrape_nsf()
        all_grants.extend(results)
        log_scan("nsf.gov", len(results), 0, 0)
        summary["sources_scanned"] += 1
    except Exception as e:
        logger.error(f"NSF scraper failed: {e}")
        log_scan("nsf.gov", 0, 0, 0, "error", str(e))
        summary["errors"].append(f"nsf.gov: {e}")

    try:
        results = scrape_nih_reporter()
        all_grants.extend(results)
        log_scan("nih_reporter", len(results), 0, 0)
        summary["sources_scanned"] += 1
    except Exception as e:
        logger.error(f"NIH Reporter scraper failed: {e}")
        log_scan("nih_reporter", 0, 0, 0, "error", str(e))
        summary["errors"].append(f"nih_reporter: {e}")

    for src in ADDITIONAL_RSS_SOURCES:
        try:
            results = scrape_rss_source(src["url"], src["name"])
            all_grants.extend(results)
            log_scan(src["name"], len(results), 0, 0)
            summary["sources_scanned"] += 1
        except Exception as e:
            logger.error(f"{src['name']} scraper failed: {e}")
            log_scan(src["name"], 0, 0, 0, "error", str(e))
            summary["errors"].append(f"{src['name']}: {e}")

    summary["total_found"] = len(all_grants)
    for grant in all_grants:
        result = upsert_grant(grant)
        summary[result] = summary.get(result, 0) + 1

    logger.info("=" * 60)
    logger.info(f"SCAN COMPLETE: {summary}")
    logger.info("=" * 60)
    return summary


# ==============================================================
# REPORTS
# ==============================================================

def get_grants_df(since_days: int = 7, source_filter: str = None) -> pd.DataFrame:
    cutoff = (utc_now() - timedelta(days=since_days)).strftime("%Y-%m-%dT%H:%M:%S")
    with get_db() as conn:
        df = pd.read_sql_query("""
            SELECT title, url, source, description, deadline,
                   posted_date, funding_amount, matched_keywords,
                   first_seen, last_seen,
                   CASE WHEN first_seen >= ? THEN 'NEW' ELSE 'UPDATED' END as status
            FROM grants
            WHERE first_seen >= ? OR (last_seen >= ? AND notified = 0)
            ORDER BY first_seen DESC
        """, conn, params=(cutoff, cutoff, cutoff))
    if not df.empty:
        df["matched_keywords"] = df["matched_keywords"].apply(
            lambda x: ", ".join(json.loads(x)) if x else ""
        )
    if source_filter and source_filter != "all":
        df = df[df["source"] == source_filter]
    return df


def get_all_grants_df() -> pd.DataFrame:
    with get_db() as conn:
        df = pd.read_sql_query("""
            SELECT title, url, source, description, deadline,
                   posted_date, funding_amount, matched_keywords,
                   first_seen, last_seen
            FROM grants
            ORDER BY
                CASE WHEN deadline IS NOT NULL AND deadline != '' THEN 0 ELSE 1 END,
                deadline ASC
        """, conn)
    if not df.empty:
        df["matched_keywords"] = df["matched_keywords"].apply(
            lambda x: json.loads(x) if x else []
        )
    return df


def export_csv(since_days: int = 7) -> str:
    df = get_grants_df(since_days)
    filename = f"grant_report_{utc_now().strftime('%Y%m%d')}.csv"
    df.to_csv(filename, index=False)
    logger.info(f"CSV exported: {filename}")
    return filename


# ==============================================================
# EXCEL EXPORT — FULL TWO-SHEET VERSION
# ==============================================================

def parse_deadline(raw: str) -> Optional[datetime]:
    if not raw or raw.strip() == "":
        return None
    formats = [
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%m/%d/%Y",
        "%B %d, %Y", "%b %d, %Y", "%m-%d-%Y", "%Y%m%d",
    ]
    cleaned = raw.strip().split("+")[0].split("Z")[0]
    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    return None


def infer_grant_type(grant_row: dict) -> str:
    deadline = str(grant_row.get("deadline", "") or "")
    posted = str(grant_row.get("posted_date", "") or "")
    title = (str(grant_row.get("title", "")) + " " + str(grant_row.get("description", ""))).lower()
    now = utc_now()

    if any(term in title for term in ["letter of intent", "loi", "pre-application"]):
        return "LOI"
    if any(term in title for term in ["rolling", "continuous", "ongoing", "no deadline"]):
        return "Ongoing"

    dl = parse_deadline(deadline)
    if dl:
        days_until = (dl - now).days
        if days_until < 0:
            return "Monitoring"
        elif days_until > 180:
            return "Future"
        else:
            return "Deadline"

    if any(term in title for term in ["forecast", "anticipated", "expected", "upcoming"]):
        return "Future"
    if posted and not deadline:
        return "Monitoring"
    return "Monitoring"


def infer_confidence(grant_row: dict) -> str:
    deadline = str(grant_row.get("deadline", "") or "")
    title = (str(grant_row.get("title", "")) + " " + str(grant_row.get("description", ""))).lower()

    if any(term in title for term in ["rolling", "continuous"]):
        return "Rolling"
    if any(term in title for term in ["annual", "yearly", "recurring", "periodic"]):
        return "Periodic"
    if any(term in title for term in ["forecast", "anticipated", "expected"]):
        return "Future"

    dl = parse_deadline(deadline)
    if dl:
        return "Confirmed"

    if any(term in title for term in ["estimated", "approximate", "tentative"]):
        return "EST"
    return "Likely"


def extract_funder(source: str, title: str, description: str) -> str:
    source_map = {
        "grants.gov": "Federal (see listing)",
        "nsf.gov": "National Science Foundation",
        "nih_reporter": "National Institutes of Health",
        "ed.gov": "U.S. Department of Education",
        "sbir.gov": "SBIR/STTR Program",
        "darpa.mil": "DARPA",
    }
    funder = source_map.get(source, "")
    combined = f"{title} {description}".lower()
    agency_patterns = {
        "national science foundation": "National Science Foundation (NSF)",
        "nsf": "National Science Foundation (NSF)",
        "department of defense": "Department of Defense (DoD)",
        "dod": "Department of Defense (DoD)",
        "department of energy": "Department of Energy (DOE)",
        "doe": "Department of Energy (DOE)",
        "department of education": "U.S. Department of Education",
        "national institutes of health": "National Institutes of Health (NIH)",
        "nih": "National Institutes of Health (NIH)",
        "darpa": "DARPA",
        "nasa": "NASA",
        "nist": "National Institute of Standards and Technology (NIST)",
        "department of transportation": "Department of Transportation (DOT)",
    }
    for pattern, agency in agency_patterns.items():
        if pattern in combined:
            return agency
    return funder if funder else "See listing"


def generate_grant_excel(output_filename: str = None) -> str:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. pip install openpyxl")
    if not HAS_DATEUTIL:
        raise ImportError("python-dateutil is required. pip install python-dateutil")
    if output_filename is None:
        output_filename = f"Grant_Tracker_{utc_now().strftime('%Y%m%d')}.xlsx"

    df = get_all_grants_df()
    if df.empty:
        return None

    wb = openpyxl.Workbook()

    # ── Styling constants ──
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    THIN_BORDER = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    TYPE_COLORS = {
        "LOI": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Deadline": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "Opens": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
        "Future": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "Monitoring": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
        "Ongoing": PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"),
    }
    CONFIDENCE_COLORS = {
        "Confirmed": Font(color="2E7D32", bold=True),
        "EST": Font(color="F57F17"),
        "Future": Font(color="1565C0"),
        "Likely": Font(color="6A1B9A"),
        "Periodic": Font(color="00695C"),
        "Rolling": Font(color="BF360C"),
    }
    CAL_DEADLINE = PatternFill(start_color="E53935", end_color="E53935", fill_type="solid")
    CAL_LOI = PatternFill(start_color="FFB300", end_color="FFB300", fill_type="solid")

    # ==========================================================
    # SHEET 1: GRANT TRACKER
    # ==========================================================

    ws1 = wb.active
    ws1.title = "Grant Tracker"
    ws1.sheet_properties.tabColor = "1F4E79"
    ws1.freeze_panes = "A2"

    tracker_columns = [
        ("Deadline Date", 16), ("Type", 14), ("Focus Area", 28),
        ("Grant / Program", 50), ("Funder", 32), ("Estimated Amount", 20),
        ("Confidence", 14), ("Action Required", 35), ("URL", 45),
    ]

    for col_idx, (header, width) in enumerate(tracker_columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    ws1.row_dimensions[1].height = 30

    for row_idx, (_, grant) in enumerate(df.iterrows(), 2):
        grant_dict = grant.to_dict()
        kw_list = grant["matched_keywords"] if isinstance(grant["matched_keywords"], list) else []

        # Deadline Date
        dl = parse_deadline(str(grant.get("deadline", "") or ""))
        deadline_cell = ws1.cell(row=row_idx, column=1)
        if dl:
            deadline_cell.value = dl
            deadline_cell.number_format = "MMM DD, YYYY"
            days_left = (dl - utc_now()).days
            if 0 < days_left <= 30:
                deadline_cell.font = Font(color="C62828", bold=True)
            elif 0 < days_left <= 60:
                deadline_cell.font = Font(color="E65100")
        else:
            deadline_cell.value = "TBD"
            deadline_cell.font = Font(color="9E9E9E", italic=True)
        deadline_cell.alignment = Alignment(horizontal="center")
        deadline_cell.border = THIN_BORDER

        # Type
        grant_type = infer_grant_type(grant_dict)
        type_cell = ws1.cell(row=row_idx, column=2, value=grant_type)
        type_cell.alignment = Alignment(horizontal="center")
        type_cell.fill = TYPE_COLORS.get(grant_type, PatternFill())
        type_cell.border = THIN_BORDER

        # Focus Area
        focus = ", ".join(kw_list)
        focus_cell = ws1.cell(row=row_idx, column=3, value=focus)
        focus_cell.alignment = Alignment(wrap_text=True, vertical="top")
        focus_cell.border = THIN_BORDER

        # Grant / Program
        title_cell = ws1.cell(row=row_idx, column=4, value=grant["title"])
        title_cell.alignment = Alignment(wrap_text=True, vertical="top")
        title_cell.font = Font(name="Calibri", size=10)
        title_cell.border = THIN_BORDER

        # Funder
        funder = extract_funder(grant["source"], grant["title"],
                                str(grant.get("description", "") or ""))
        funder_cell = ws1.cell(row=row_idx, column=5, value=funder)
        funder_cell.alignment = Alignment(wrap_text=True, vertical="top")
        funder_cell.border = THIN_BORDER

        # Estimated Amount
        amount_cell = ws1.cell(row=row_idx, column=6)
        raw_amount = str(grant.get("funding_amount", "") or "").strip()
        try:
            numeric = float(re.sub(r'[^\d.]', '', raw_amount))
            if numeric > 0:
                amount_cell.value = numeric
                amount_cell.number_format = '$#,##0'
            else:
                amount_cell.value = "See listing"
                amount_cell.font = Font(color="9E9E9E", italic=True)
        except (ValueError, TypeError):
            amount_cell.value = raw_amount if raw_amount else "See listing"
            if not raw_amount:
                amount_cell.font = Font(color="9E9E9E", italic=True)
        amount_cell.alignment = Alignment(horizontal="center")
        amount_cell.border = THIN_BORDER

        # Confidence
        confidence = infer_confidence(grant_dict)
        conf_cell = ws1.cell(row=row_idx, column=7, value=confidence)
        conf_cell.font = CONFIDENCE_COLORS.get(confidence, Font())
        conf_cell.alignment = Alignment(horizontal="center")
        conf_cell.border = THIN_BORDER

        # Action Required
        action = ""
        if grant_type == "Deadline":
            if dl:
                days_left = (dl - utc_now()).days
                if days_left <= 14:
                    action = "URGENT: Submit application"
                elif days_left <= 30:
                    action = "Finalize application materials"
                elif days_left <= 60:
                    action = "Begin application drafting"
                else:
                    action = "Review requirements; plan timeline"
            else:
                action = "Review requirements"
        elif grant_type == "LOI":
            action = "Prepare Letter of Intent"
        elif grant_type == "Future":
            action = "Monitor for announcement"
        elif grant_type == "Monitoring":
            action = "Watch for next cycle"
        elif grant_type == "Ongoing":
            action = "Review eligibility; apply when ready"
        else:
            action = "Review opportunity"
        action_cell = ws1.cell(row=row_idx, column=8, value=action)
        action_cell.alignment = Alignment(wrap_text=True, vertical="top")
        action_cell.border = THIN_BORDER

        # URL
        url_cell = ws1.cell(row=row_idx, column=9, value=grant["url"])
        url_cell.font = Font(color="0563C1", underline="single", size=9)
        url_cell.hyperlink = grant["url"]
        url_cell.alignment = Alignment(wrap_text=True, vertical="top")
        url_cell.border = THIN_BORDER

        ws1.row_dimensions[row_idx].height = 35

    # Autofilter
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(tracker_columns))}{len(df) + 1}"

    # Data validation dropdowns
    type_validation = DataValidation(
        type="list",
        formula1='"LOI,Deadline,Opens,Future,Monitoring,Ongoing"',
        allow_blank=True,
    )
    type_validation.error = "Please select a valid type"
    ws1.add_data_validation(type_validation)
    type_validation.add(f"B2:B{len(df) + 1}")

    conf_validation = DataValidation(
        type="list",
        formula1='"Confirmed,EST,Future,Likely,Periodic,Rolling"',
        allow_blank=True,
    )
    ws1.add_data_validation(conf_validation)
    conf_validation.add(f"G2:G{len(df) + 1}")

    # ==========================================================
    # SHEET 2: CALENDAR VIEW
    # ==========================================================

    ws2 = wb.create_sheet("Calendar View")
    ws2.sheet_properties.tabColor = "2E7D32"
    ws2.freeze_panes = "C2"

    now = utc_now()
    months = []
    for i in range(24):
        months.append(now + relativedelta(months=i))

    cal_columns = ["Focus Area", "Grant / Program"]
    for m in months:
        cal_columns.append(m.strftime("%b %Y"))
    cal_columns.append("Notes / Key Actions")

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 42
    for col_idx in range(3, 3 + 24):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 12
    ws2.column_dimensions[get_column_letter(27)].width = 35

    for col_idx, header in enumerate(cal_columns, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws2.row_dimensions[1].height = 30

    for row_idx, (_, grant) in enumerate(df.iterrows(), 2):
        grant_dict = grant.to_dict()
        kw_list = grant["matched_keywords"] if isinstance(grant["matched_keywords"], list) else []

        # Focus Area
        focus = ", ".join(kw_list)
        focus_cell = ws2.cell(row=row_idx, column=1, value=focus)
        focus_cell.alignment = Alignment(wrap_text=True, vertical="center")
        focus_cell.border = THIN_BORDER

        # Grant / Program
        title_cell = ws2.cell(row=row_idx, column=2, value=grant["title"])
        title_cell.alignment = Alignment(wrap_text=True, vertical="center")
        title_cell.font = Font(name="Calibri", size=9)
        title_cell.border = THIN_BORDER

        # Month columns
        dl = parse_deadline(str(grant.get("deadline", "") or ""))
        grant_type = infer_grant_type(grant_dict)

        for month_idx, month_date in enumerate(months):
            col = month_idx + 3
            cell = ws2.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if dl:
                if dl.year == month_date.year and dl.month == month_date.month:
                    if grant_type == "LOI":
                        cell.value = "LOI Due"
                        cell.fill = CAL_LOI
                        cell.font = Font(bold=True, size=9, color="FFFFFF")
                    else:
                        cell.value = "DEADLINE"
                        cell.fill = CAL_DEADLINE
                        cell.font = Font(bold=True, size=9, color="FFFFFF")

                prep_date = dl - relativedelta(months=1)
                if prep_date.year == month_date.year and prep_date.month == month_date.month:
                    cell.value = "Prepare"
                    cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    cell.font = Font(size=9, color="F57F17")

                plan_date = dl - relativedelta(months=2)
                if plan_date.year == month_date.year and plan_date.month == month_date.month:
                    cell.value = "Plan"
                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    cell.font = Font(size=9, color="1565C0")

            elif grant_type == "Ongoing":
                cell.value = "Open"
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                cell.font = Font(size=8, color="2E7D32")

            elif grant_type == "Monitoring":
                if month_idx < 6:
                    cell.value = "Watch"
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    cell.font = Font(size=8, color="9E9E9E")

        # Notes / Key Actions
        notes = ""
        if dl:
            days_left = (dl - utc_now()).days
            if days_left > 0:
                notes = f"{days_left} days until deadline"
            elif days_left == 0:
                notes = "DEADLINE TODAY"
            else:
                notes = f"Closed {abs(days_left)} days ago - watch for next cycle"
        elif grant_type == "Ongoing":
            notes = "Rolling - apply when ready"
        else:
            notes = "Monitor for updates"

        notes_cell = ws2.cell(row=row_idx, column=27, value=notes)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="center")
        notes_cell.border = THIN_BORDER

        ws2.row_dimensions[row_idx].height = 30

    # Legend
    legend_start = len(df) + 3
    ws2.cell(row=legend_start, column=1, value="LEGEND").font = Font(bold=True, size=11)

    legend_items = [
        ("DEADLINE", CAL_DEADLINE, Font(bold=True, color="FFFFFF", size=9)),
        ("LOI Due", CAL_LOI, Font(bold=True, color="FFFFFF", size=9)),
        ("Prepare", PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
         Font(size=9, color="F57F17")),
        ("Plan", PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
         Font(size=9, color="1565C0")),
        ("Open (Rolling)", PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
         Font(size=8, color="2E7D32")),
        ("Watch", PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
         Font(size=8, color="9E9E9E")),
    ]

    for i, (label, fill, font) in enumerate(legend_items):
        row = legend_start + 1 + i
        sample = ws2.cell(row=row, column=1, value=label)
        sample.fill = fill
        sample.font = font
        sample.alignment = Alignment(horizontal="center")
        sample.border = THIN_BORDER
        ws2.cell(row=row, column=2, value=f"= {label}").font = Font(size=9, color="666666")

    # Save
    wb.save(output_filename)
    logger.info(f"Excel workbook saved: {output_filename}")

    file_size = Path(output_filename).stat().st_size / 1024
    logger.info(f"  Sheet 1 - Grant Tracker: {len(df)} grants")
    logger.info(f"  Sheet 2 - Calendar View: {len(df)} grants x 24 months")
    logger.info(f"  File size: {file_size:.1f} KB")

    return output_filename
